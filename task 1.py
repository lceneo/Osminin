import csv
import datetime
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from builtins import input
import re
from os import listdir
from cProfile import Profile
from multiprocessing import Process
from pstats import Stats


def multi_proccessing(f_name, my_inputter):
    info_set = DataSet(f_name, list())
    info_set.fill_vacancies()
    my_inputter.count_vacancies(info_set.vacancies_objects)

class Vacancy:
    """
    Класс для представления вакансий.

    Attributes:
       name (str): название вакансии
       salary (int): зарплата
       area_name (str): область деятельности
       published_at (datetime): время публикации вакансии
    """
    currency_to_rub = {
        "KGS": 0.76,
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, object_vacancy):
        """
            Инициализирует объект класса Vacancy, инициалазируя значениями поля класса

            Args:
                object_vacancy(obj): объект, представляющий основную информацию о вакансии
        """

        self.name = object_vacancy['name']
        salary_from = (int)((float)("".join(object_vacancy['salary_from'].split())))
        salary_to = (int)((float)("".join(object_vacancy['salary_to'].split())))
        self.salary = (salary_from + salary_to) * self.currency_to_rub[object_vacancy['salary_currency']] // 2
        self.area_name = object_vacancy['area_name']
        #self.published_at = datetime.datetime.strptime(object_vacancy['published_at'], '%Y-%m-%dT%H:%M:%S%z')
        #self.published_at = self.timearea_parse(object_vacancy['published_at'])
        self.published_at = self.datetime_parse(object_vacancy['published_at'])
    
    def datetime_parse(self, date: str):
        return datetime.datetime(int(date[0:4]), int(date[5:7]), int(date[8:10]), int(date[11:13]),
                                 int(date[14:16]), int(date[17:19]), int(date[19:22]))
    
    def timearea_parse(self, date: str):
        hms = [int(key) for key in (date.split('T')[1]).split('+')[0].split(':')]
        ydm = [int(key) for key in (date.split('T')[0]).split('-')]
        timearea = int(date[19:22])
        return datetime.datetime(ydm[0], ydm[1], ydm[2], hms[0], hms[1], hms[2], timearea)

class DataSet:
    """
    Класс для формирования набора данных по вакансиям
    Attributes:
        file_name(str): название файла с вакансиями
        vacancies_objects(list): список вакансий
    """

    def __init__(self, file_name: str, vacancies_objects: list):
        """
        Инициализирует объект DataSet под указанным именем и с переданным набором данных
        Args:
            file_name(str): название файла
            vacancies_objects(list): список вакансий
        """
        self.file_name = file_name
        self.vacancies_objects = vacancies_objects

    def refactor_html(self, raw_html):
        """
        Очищает текст от html-тегов
        Args:
            raw_html: html-разметка
        Returns:
            str: Очищенный от html-тегов текст
        """
        return (re.compile('<.*?>'), '', raw_html)

    def csv_read(self):
        """
        Считывает данные из переданного файла
        Returns:
            Tuple: пара значений(вакансия - имя вакансии)
          """
        vacancies = []
        name_list = []
        with open(self.file_name, encoding='utf-8-sig') as r_file:
            file_reader = csv.reader(r_file, delimiter=",")
            count = 0
            for row in file_reader:
                if count == 0:
                    count += 1
                    name_list = row
                else:
                    if "" in row or len(row) != len(name_list):
                        continue
                    vacancies.append(row)
        if len(name_list) == 0:
            print('Пустой файл')
            exit()
        if len(vacancies) == 0:
            print('Нет данных')
            exit()
        return (vacancies, name_list)

    def fill_vacancies(self):
        """
        Заполняет объекты вакансий полученной информацией
        """
        (vacancies, list_naiming) = self.csv_read()
        self.vacancies_objects = self.csv_filer(vacancies, list_naiming)

    def csv_filer(self, reader, list_naming):
        """
        Идёт построчно по переданному файлу и заполняет вакансии
        Args:
            reader -
            list_naming(list) - список названий вакансий
        Returns:
            list: Список сформированных факансий
        """
        vacancies = list()
        for row in reader:
            current = {}
            for i in range(len(row)):
                current[list_naming[i]] = row[i]
            vacancies.append(Vacancy(current))
        return vacancies


class Tuple:
    """
    Класс для хранения двух параметров
    Attributes:
        totalSalary(int): финальная зарплата
        count(int): номер вакансии
    """
    totalSalary = 0
    count = 0

    def __init__(self, totalSalary: int, count: int):
        self.totalSalary = totalSalary
        self.count = count


class InputData:
    """
    Класс для формирования статистики по вакансиям(изменения популярности по годам, городам, профессиям)
    Attributes:
        years_stats(dict): статистика по годам
        cities_stats(dict): статистика по городам
        vacancy_stats(dict): статистика по профессиям
    """

    years_stats = {
    }

    cities_stats = {
    }

    vacancy_stats = {
    }

    def start_input(self):
        """
        Инициализирование данных по новой профессии
        """
        self.file_name = input('Введите название файла: ')
        self.profession = input('Введите название профессии: ')
        self.city_count = 0

    def count_vacancies(self, vacancies: list):
        """
        Подсчёт данных по городам/годам/профессиям
        Args:
            vacancies(list): список вакансий
        """
        for vacancy in vacancies:
            self.city_count += 1
            year = int(vacancy.published_at.year)

            if vacancy.area_name not in self.cities_stats.keys():
                self.cities_stats[vacancy.area_name] = Tuple(vacancy.salary, 1)
            else:
                self.cities_stats[vacancy.area_name].totalSalary += vacancy.salary
                self.cities_stats[vacancy.area_name].count += 1

            if year not in self.years_stats.keys():
                self.years_stats[year] = Tuple(vacancy.salary, 1)
                self.vacancy_stats[year] = Tuple(0, 0)
            else:
                self.years_stats[year].totalSalary += vacancy.salary
                self.years_stats[year].count += 1

            if self.profession in vacancy.name:
                self.vacancy_stats[year].totalSalary += vacancy.salary
                self.vacancy_stats[year].count += 1

    def update_stats(self):
        """
        Обновление данных по городам/годам/профессиям
        """
        for year in self.years_stats.keys():
            self.years_stats[year].totalSalary = int(self.years_stats[year].totalSalary // self.years_stats[year].count)

        remove_data = list()
        for city in self.cities_stats.keys():
            percentage = round(self.cities_stats[city].count / self.city_count, 4)
            if percentage < 0.01:
                remove_data.append(city)
            else:
                self.cities_stats[city].totalSalary = int(
                    self.cities_stats[city].totalSalary // self.cities_stats[city].count)
                self.cities_stats[city].count = percentage
        for year in self.vacancy_stats.keys():
            if self.vacancy_stats[year].count != 0:
                self.vacancy_stats[year].totalSalary = int(
                    self.vacancy_stats[year].totalSalary // self.vacancy_stats[year].count)
        for city in remove_data:
            del [self.cities_stats[city]]

    def print_info(self, str_info: str, dict: dict, value_name: str):
        """
        Вывод информации-статистики по годам
        Args:
            str_info(str): краткая информация по вакансии
            dict(dict): словарь вакансий
            value_name(str): имя вакансии
        """
        marker = False
        print(str_info, end='')
        ind = 0
        for year in dict.keys():
            if ind == 0:
                print(' {', end='')
                ind += 1
            printEnd = ', '
            if year == max(dict.keys()):
                printEnd = ''
                marker = True
            print(f"{year}: {getattr(dict[year], value_name)}", end=printEnd)
        if marker:
            print('}')

    def get_city_print(self, str_data: str, dict: dict, names: list, value_name):
        """
        Вывод информации-статистики по городам
        Args:
            str_info(str): краткая информация по вакансии
            dict(dict): словарь вакансий
            names(list): список названий вакансий
        """
        flag = False
        print(str_data, end='')
        ind = 0
        for name in names:
            if ind == 0:
                print(' {', end='')
            printEnd = ', '
            if ind == len(names) - 1:
                printEnd = ''
                flag = True
            print(f"'{name}': {getattr(dict[name], value_name)}", end=printEnd)
            ind += 1
        if flag:
            print('}')

    def get_answer(self):
        """
        Вывод всей сформированной информации
        """

        cities_sorted = sorted(self.cities_stats, key=lambda x: self.cities_stats[x].totalSalary, reverse=True)
        self.print_info("Динамика уровня зарплат по годам:", self.years_stats, "totalSalary")
        self.print_info("Динамика уровня зарплат по годам для выбранной профессии:", self.vacancy_stats, "totalSalary")

        self.print_info("Динамика количества вакансий по годам:", self.years_stats, "count")

        self.print_info("Динамика количества вакансий по годам для выбранной профессии:", self.vacancy_stats, "count")
        del cities_sorted[10:]
        self.get_city_print("Уровень зарплат по городам (в порядке убывания):", self.cities_stats,
                            cities_sorted, "totalSalary")
        cities_sorted = sorted(self.cities_stats, key=lambda x: self.cities_stats[x].count, reverse=True)
        del cities_sorted[10:]
        self.get_city_print("Доля вакансий по городам (в порядке убывания):", self.cities_stats,
                            cities_sorted, "count")

    def get_sorted_cities(self, attr_name: str):
        """
        Сортирует вакансии по городам
        Args:
            attr_name(str): имя города для сортировки
        Returns:
            dic: словарь с вакансиями по городам
        """
        current = {}
        sorted_names = sorted(self.cities_stats, key=lambda x: getattr(self.cities_stats[x], attr_name), reverse=True)
        del sorted_names[10:]
        for name in sorted_names:
            current[name] = self.cities_stats[name]
        return current


class Report:
    """
    Класс для формирования отчёта по вакансиям
    """

    def generate_excel(self, inputer: InputData):
        """
        Формирует отчёт в формате Excel по полученным данным
        Args:
            inputer(InputData): данные для формирования отчёта
        """
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        wb = openpyxl.Workbook()
        wb.remove(wb['Sheet'])
        wb.create_sheet("Статистика по годам")
        list = wb["Статистика по годам"]
        list['A1'].font = Font(bold=True)
        list['A1'].border = thin_border
        list['A1'] = "Год"
        list.column_dimensions['A'].width = 6
        list['B1'] = "Средняя зарплата"
        list['B1'].font = Font(bold=True)
        list['B1'].border = thin_border
        list.column_dimensions['B'].width = len("Средняя зарплата") + 2
        list['C1'] = f"Средняя зарплата - {inputer.profession}"
        list['C1'].font = Font(bold=True)
        list['C1'].border = thin_border
        list.column_dimensions['C'].width = len(f"Средняя зарплата - {inputer.profession}") + 2
        list['D1'] = "Количество вакансий"
        list['D1'].font = Font(bold=True)
        list['D1'].border = thin_border
        list.column_dimensions['D'].width = len("Количество вакансий") + 2
        list['E1'].font = Font(bold=True)
        list['E1'].border = thin_border
        list['E1'] = f"Количество вакансий - {inputer.profession}"
        list.column_dimensions['E'].width = len(f"Количество вакансий - {inputer.profession}") + 2
        for i in inputer.years_stats:
            list[f"A{i - 2005}"] = i
            list[f"A{i - 2005}"].border = thin_border
            list[f"B{i - 2005}"] = inputer.years_stats[i].totalSalary
            list[f"B{i - 2005}"].border = thin_border
            list[f"D{i - 2005}"] = inputer.years_stats[i].count
            list[f"D{i - 2005}"].border = thin_border
        for i in inputer.vacancy_stats:
            list[f"C{i - 2005}"] = inputer.vacancy_stats[i].totalSalary
            list[f"C{i - 2005}"].border = thin_border
            list[f"E{i - 2005}"] = inputer.vacancy_stats[i].count
            list[f"E{i - 2005}"].border = thin_border

        wb.create_sheet("Статистика по городам")
        list = wb["Статистика по городам"]
        list['A1'] = "Город"
        list['A1'].font = Font(bold=True)
        list['A1'].border = thin_border
        list.column_dimensions['A'].width = len("Город") + 2
        list['B1'] = "Уровень зарплат"
        list['B1'].font = Font(bold=True)
        list['B1'].border = thin_border
        list.column_dimensions['B'].width = len("Уровень зарплат") + 2
        list['D1'] = "Город"
        list['D1'].font = Font(bold=True)
        list['D1'].border = thin_border
        list.column_dimensions['D'].width = len("Город") + 2
        list['E1'] = "Доля ваканский"
        list['E1'].font = Font(bold=True)
        list['E1'].border = thin_border
        list.column_dimensions['E'].width = len("Доля ваканский") + 2
        sorted_cities = inputer.get_sorted_cities("totalSalary")
        ind = 2
        for i in sorted_cities:
            list[f"A{ind}"] = i
            list[f"A{ind}"].border = thin_border
            list.column_dimensions['A'].width = max(list.column_dimensions['A'].width, len(i) + 2)
            list[f"B{ind}"] = sorted_cities[i].totalSalary
            list[f"B{ind}"].border = thin_border
            ind = ind + 1
        sorted_cities = inputer.get_sorted_cities("count")
        ind = 2
        for i in sorted_cities:
            list[f"D{ind}"] = i
            list[f"D{ind}"].border = thin_border
            list.column_dimensions['D'].width = max(list.column_dimensions['D'].width, len(i) + 2)
            list[f"E{ind}"] = f"{round(sorted_cities[i].count * 100, 2)}%"
            list[f"E{ind}"].number_format = '0.00%'
            list[f"E{ind}"].border = thin_border
            ind = ind + 1
        wb.save("report.xlsx")
        
class Tests(TestCase):

    def test_clean_html(self):

        self.assertEqual(DataSet("", list()).cleanhtml("<tag>text</tag>"), 'text')
    def test_salary1(self):
        self.assertEqual(Vacancy({"name" : "","salary_from" : 12, "salary_to" : 35, "salary_currency" : "RUR", 
                                  "area_name" : "", "published_at" : datetime.datetime(3,6,7,8,9,10,4)}).salary_from, 12);
        
    def test_salary2(self):
        self.assertEqual(Vacancy({"name" : "","salary_from" : 12, "salary_to" : 35, "salary_currency" : "RUR", 
                                  "area_name" : "", "published_at" : datetime.datetime(3,6,7,8,9,10,4)}).salary_to, 35);

    def test_tuple_2(self):
        self.assertEqual(Tuple(0, 42).count, 42)       
                   
    def test_tuple1(self):
        self.assertEqual(Tuple(3210, 1).totalSalary, 3210)
        
 
inputer = InputData()
inputer.start_input()
dataset = DataSet(inputer.file_name, list())
dataset.fill_vacancies()
inputer.count_vacancies(dataset.vacancies_objects)
inputer.update_stats()
inputer.get_answer()
report = Report()
report.generate_excel(inputer)
profile = Profile()
inputer = InputData()
profile.disable()
inputer.start_input()
profile.enable()
multiP = Process(target=multi_proccessing, args=(listdir(f'./{inputer.file_name}'),inputer))
multiP.start()
multiP.join()
dataset = DataSet(inputer.file_name, list())
dataset.fill_vacancies()
inputer.count_vacancies(dataset.vacancies_objects)
inputer.update_stats()
inputer.get_answer()
profile.disable()
profile.dump_stats('mystats.stats')
with open('mystats_output.txt', 'wt') as output:
    stats = Stats('mystats.stats', stream=output)
    stats.sort_stats('cumulative', 'time')
    stats.print_stats()
